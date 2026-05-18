# -*- coding: utf-8 -*-
from datetime import datetime
from pathlib import Path
import csv
import re

from openpyxl import load_workbook


SOURCE = Path("gsc-source.xlsx")
OUT = Path("gsc-top100-report.tsv")


def clean_pos(value):
    if isinstance(value, datetime):
        return f"{value.day}.{value.month:02d}"
    if value is None:
        return ""
    return str(value)


def n(value):
    if value in (None, ""):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def s(value):
    return "" if value is None else str(value).strip()


def has(q, pattern):
    return re.search(pattern, q, re.I) is not None


def classify(query, clicks, impressions):
    q = query.lower()
    if has(q, r"kna|knacert"):
        group = "Brand KNA"
    elif has(q, r"iso 9001|iso 14001|iso 22000|iso 45001|iso 13485|iso 17025|iso 27001|iso 50001|iso 31000|iso 3834|\biso\b"):
        group = "Chứng nhận ISO"
    elif has(q, r"haccp|fssc|gmp|gfsi|brc|halal|kosher|vietgap|globalgap|ssop|oprp|ccp"):
        group = "An toàn thực phẩm"
    elif has(q, r"bsci|smeta|sa ?8000|rba|sedex|wrap|wca|ctpat|fla|better work|higg|oeko|grs|rcs"):
        group = "Audit/trách nhiệm xã hội"
    elif has(q, r"fda|\bce\b|ce marking|cfs|duns|ul|gacc|eudr|fsc|pefc|msc|iscc|fair trade|fairtrade"):
        group = "Chứng nhận xuất khẩu"
    elif has(q, r"7 ?qc|7qc|fmea|spc|msa|kaizen|lean|six sigma|5s|kanban|swot|poka|pdca|qcc"):
        group = "Công cụ cải tiến/quản lý"
    else:
        group = "Khác"

    standards = [
        ("ISO 9001", r"iso 9001"),
        ("ISO 14001", r"iso 14001"),
        ("ISO 22000/FSSC 22000", r"iso 22000|fssc"),
        ("ISO 45001", r"iso 45001"),
        ("ISO 13485", r"iso 13485"),
        ("ISO 17025", r"iso 17025|17025"),
        ("ISO 27001", r"iso 27001"),
        ("HACCP", r"haccp"),
        ("GMP/cGMP", r"gmp|cgmp"),
        ("BSCI", r"bsci"),
        ("SMETA/Sedex", r"smeta|sedex"),
        ("SA8000", r"sa ?8000"),
        ("IATF 16949", r"iatf"),
        ("FDA", r"fda"),
        ("CE", r"\bce\b|ce marking"),
        ("FSC", r"fsc"),
        ("PEFC", r"pefc"),
        ("WRAP", r"wrap"),
    ]
    standard = next((name for name, pat in standards if has(q, pat)), "Khác")

    if has(q, r"thực phẩm|haccp|fssc|gmp|brc|halal|kosher|vietgap|globalgap|ssop|oprp|ccp"):
        industry = "Thực phẩm"
    elif has(q, r"gỗ|wood|eudr|fsc|pefc|viên nén"):
        industry = "Gỗ/xuất khẩu"
    elif has(q, r"dệt|may|oeko|textile|grs|rcs|wrap"):
        industry = "Dệt may"
    elif has(q, r"thiết bị y tế|medical|13485"):
        industry = "Thiết bị y tế"
    elif has(q, r"iatf|ô tô|automotive"):
        industry = "Ô tô"
    else:
        industry = "Đa ngành"

    if has(q, r"kna|knacert|công ty tnhh chứng nhận kna"):
        intent = "Điều hướng"
    elif has(q, r"học|khóa học|đào tạo|dịch vụ|đăng ký|thủ tục|cấp chứng nhận|check|tra cứu"):
        intent = "Giao dịch"
    elif has(q, r"chi phí|giá|báo giá|tổ chức chứng nhận|công ty chứng nhận|uy tín|chứng nhận"):
        intent = "Thương mại"
    else:
        intent = "Thông tin"

    funnel = {"Giao dịch": "BOFU", "Thương mại": "MOFU", "Điều hướng": "BOFU"}.get(intent, "TOFU")
    page_type = {
        "Giao dịch": "Landing page/Trang dịch vụ",
        "Thương mại": "Trang dịch vụ/So sánh",
        "Điều hướng": "Trang brand",
    }.get(intent, "Bài blog/Pillar")
    if has(q, r"pdf|mẫu|form|checklist"):
        page_type = "Tài nguyên/Download"

    cluster = standard if standard != "Khác" else group
    if impressions >= 1000:
        opportunity = "Nhiều hiển thị - cần tối ưu CTR/title"
    elif clicks >= 10:
        opportunity = "Đang có click - nên tối ưu để tăng lead"
    elif impressions >= 100:
        opportunity = "Có nhu cầu tìm kiếm - cần rà intent"
    else:
        opportunity = "Long-tail/ưu tiên sau"

    if intent in ("Giao dịch", "Thương mại"):
        priority = "Cao"
    elif impressions >= 1000 or clicks >= 10:
        priority = "Trung bình"
    else:
        priority = "Thấp"

    if has(q, r"pdf|download|mẫu|form|checklist"):
        note = "Có thể tạo lead magnet hoặc tài nguyên tải về"
    elif has(q, r"kna|knacert"):
        note = "Brand query - kiểm tra trang đích và CTR"
    elif has(q, r"là gì|tiêu chuẩn|quy trình|hướng dẫn"):
        note = "Keyword thông tin - dùng internal link về trang dịch vụ"
    else:
        note = "Cần kiểm tra SERP trước khi quyết định page"

    return group, standard, industry, intent, funnel, page_type, cluster, opportunity, priority, note


def main():
    wb = load_workbook(SOURCE, data_only=True)
    ws = wb.worksheets[0]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=101, values_only=True):
        query = s(row[0])
        if not query:
            continue
        clicks = n(row[1])
        impressions = n(row[2])
        values = [
            query,
            clicks,
            impressions,
            s(row[3]),
            clean_pos(row[4]),
            query.lower(),
            *classify(query, clicks, impressions),
        ]
        rows.append(values)

    headers = [
        "truy_van", "luot_nhap", "hien_thi", "ctr", "vi_tri", "tu_khoa_chuan_hoa",
        "nhom_dich_vu", "tieu_chuan_chung_nhan", "nganh_chu_de", "y_dinh_chinh",
        "giai_doan_pheu", "loai_trang", "cum_chu_de", "co_hoi_seo", "uu_tien_seo", "ghi_chu"
    ]
    with OUT.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        writer.writerows(rows)
    print(OUT.resolve())


if __name__ == "__main__":
    main()
