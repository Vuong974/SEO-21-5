# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

wb = openpyxl.load_workbook('Phân Loại Page KNA CERT (1).xlsx', data_only=True)
ws5 = wb['Page Tháng 5']

# Collect data
nhom_data = defaultdict(lambda: {'clicks':0,'impressions':0,'events':0,'pages':0})
loai_data = defaultdict(lambda: {'clicks':0,'impressions':0,'events':0,'pages':0})
total = {'clicks':0,'impressions':0,'events':0,'pages':0}

for r in range(2, 800):
    url = ws5.cell(r, 3).value
    if not url:
        continue
    nhom = ws5.cell(r, 6).value or 'Không xác định'
    loai = ws5.cell(r, 5).value or 'Không xác định'
    click = ws5.cell(r, 12).value or 0
    hienthi = ws5.cell(r, 13).value or 0
    sukien = ws5.cell(r, 14).value or 0
    nhom_data[nhom]['clicks'] += click
    nhom_data[nhom]['impressions'] += hienthi
    nhom_data[nhom]['events'] += sukien
    nhom_data[nhom]['pages'] += 1
    loai_data[loai]['clicks'] += click
    loai_data[loai]['impressions'] += hienthi
    loai_data[loai]['events'] += sukien
    loai_data[loai]['pages'] += 1
    total['clicks'] += click
    total['impressions'] += hienthi
    total['events'] += sukien
    total['pages'] += 1

# Create or replace sheet
if 'Báo Cáo Tháng 5' in wb.sheetnames:
    del wb['Báo Cáo Tháng 5']
ws = wb.create_sheet('Báo Cáo Tháng 5', 0)

# Style helpers
BLUE_DARK  = 'FF1F3864'
BLUE_MID   = 'FF2E75B6'
BLUE_LIGHT = 'FFD6E4F0'
GRAY_LIGHT = 'FFF2F2F2'
WHITE      = 'FFFFFFFF'
ORANGE_BG  = 'FFED7D31'
GREEN_BG   = 'FF70AD47'
ORANGE_LT  = 'FFFCE4D6'
GREEN_LT   = 'FFE2EFDA'
NAVY_TXT   = 'FFFFFFFF'

def thin_border():
    s = Side(style='thin', color='FFB8CCE4')
    return Border(left=s, right=s, top=s, bottom=s)

def cs(row, col, value, bold=False, bg=None, fg='FF000000', align='left', size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=size, name='Calibri')
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = thin_border()
    return c

# ===== TITLE =====
ws.merge_cells('A1:F1')
c = ws.cell(1, 1, 'BÁO CÁO TỔNG QUAN THÁNG 5 - KNA CERT')
c.font = Font(bold=True, size=16, color=NAVY_TXT, name='Calibri')
c.fill = PatternFill('solid', fgColor=BLUE_DARK)
c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 38

ws.merge_cells('A2:F2')
c2 = ws.cell(2, 1, f'Tổng số trang phân tích: {total["pages"]} trang | Nguồn dữ liệu: Google Search Console')
c2.font = Font(italic=True, size=10, color='FF595959', name='Calibri')
c2.fill = PatternFill('solid', fgColor='FFD9E1F2')
c2.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 18

# spacer row 3
ws.row_dimensions[3].height = 8

# ===== TONG TOAN WEBSITE =====
ws.merge_cells('A4:F4')
c3 = ws.cell(4, 1, 'TỔNG TOÀN WEBSITE')
c3.font = Font(bold=True, size=13, color=NAVY_TXT, name='Calibri')
c3.fill = PatternFill('solid', fgColor=BLUE_MID)
c3.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[4].height = 30

tot_headers = ['Số trang', 'Lượt nhấp (Click)', 'Lượt hiển thị', 'Sự kiện quan trọng', 'CTR (%)']
for i, h in enumerate(tot_headers, 1):
    cs(5, i, h, bold=True, bg=BLUE_LIGHT, align='center')
ws.row_dimensions[5].height = 22

ctr_total = round(total['clicks'] / total['impressions'] * 100, 2) if total['impressions'] else 0
tot_vals = [total['pages'], total['clicks'], total['impressions'], total['events'], f'{ctr_total}%']
for i, v in enumerate(tot_vals, 1):
    cs(6, i, v, bold=True, bg=WHITE, align='center', size=13)
ws.row_dimensions[6].height = 28

# spacer row 7
ws.row_dimensions[7].height = 8

# ===== NHOM TRANG =====
ws.merge_cells('A8:F8')
c4 = ws.cell(8, 1, 'PHÂN TÍCH THEO NHÓM TRANG')
c4.font = Font(bold=True, size=12, color=NAVY_TXT, name='Calibri')
c4.fill = PatternFill('solid', fgColor=ORANGE_BG)
c4.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[8].height = 28

nh_cols = ['Nhóm Trang', 'Số trang', 'Lượt nhấp', 'Lượt hiển thị', 'Sự kiện', 'CTR (%)']
for i, h in enumerate(nh_cols, 1):
    cs(9, i, h, bold=True, bg=ORANGE_LT, align='center')
ws.row_dimensions[9].height = 22

nhom_order = ['Blog – Tin tức', 'Service – Chứng nhận', 'Service – Tư vấn', 'Service – Đào tạo', 'Trang tĩnh']
for idx, nhom in enumerate(nhom_order):
    v = nhom_data[nhom]
    r = 10 + idx
    bg = WHITE if idx % 2 == 0 else GRAY_LIGHT
    ctr_n = round(v['clicks'] / v['impressions'] * 100, 2) if v['impressions'] else 0
    cs(r, 1, nhom, bg=bg)
    cs(r, 2, v['pages'], bg=bg, align='center')
    cs(r, 3, v['clicks'], bg=bg, align='center')
    cs(r, 4, v['impressions'], bg=bg, align='center')
    cs(r, 5, v['events'], bg=bg, align='center')
    cs(r, 6, f'{ctr_n}%', bg=bg, align='center')
    ws.row_dimensions[r].height = 22

r_nh_tot = 10 + len(nhom_order)
cs(r_nh_tot, 1, 'TỔNG', bold=True, bg=BLUE_LIGHT, align='center')
cs(r_nh_tot, 2, total['pages'], bold=True, bg=BLUE_LIGHT, align='center')
cs(r_nh_tot, 3, total['clicks'], bold=True, bg=BLUE_LIGHT, align='center')
cs(r_nh_tot, 4, total['impressions'], bold=True, bg=BLUE_LIGHT, align='center')
cs(r_nh_tot, 5, total['events'], bold=True, bg=BLUE_LIGHT, align='center')
cs(r_nh_tot, 6, f'{ctr_total}%', bold=True, bg=BLUE_LIGHT, align='center')
ws.row_dimensions[r_nh_tot].height = 22

# spacer
ws.row_dimensions[r_nh_tot + 1].height = 8

# ===== LOAI TRANG =====
r_lt = r_nh_tot + 2
ws.merge_cells(f'A{r_lt}:F{r_lt}')
c5 = ws.cell(r_lt, 1, 'PHÂN TÍCH THEO LOẠI TRANG')
c5.font = Font(bold=True, size=12, color=NAVY_TXT, name='Calibri')
c5.fill = PatternFill('solid', fgColor=GREEN_BG)
c5.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[r_lt].height = 28

r_lt_h = r_lt + 1
lt_cols = ['Loại Trang', 'Số trang', 'Lượt nhấp', 'Lượt hiển thị', 'Sự kiện', 'CTR (%)']
for i, h in enumerate(lt_cols, 1):
    cs(r_lt_h, i, h, bold=True, bg=GREEN_LT, align='center')
ws.row_dimensions[r_lt_h].height = 22

loai_order = sorted(loai_data.keys())
for idx, loai in enumerate(loai_order):
    v = loai_data[loai]
    r = r_lt_h + 1 + idx
    bg = WHITE if idx % 2 == 0 else GRAY_LIGHT
    ctr_l = round(v['clicks'] / v['impressions'] * 100, 2) if v['impressions'] else 0
    cs(r, 1, loai, bg=bg)
    cs(r, 2, v['pages'], bg=bg, align='center')
    cs(r, 3, v['clicks'], bg=bg, align='center')
    cs(r, 4, v['impressions'], bg=bg, align='center')
    cs(r, 5, v['events'], bg=bg, align='center')
    cs(r, 6, f'{ctr_l}%', bg=bg, align='center')
    ws.row_dimensions[r].height = 22

r_lt_tot = r_lt_h + 1 + len(loai_order)
cs(r_lt_tot, 1, 'TỔNG', bold=True, bg=GREEN_LT, align='center')
cs(r_lt_tot, 2, total['pages'], bold=True, bg=GREEN_LT, align='center')
cs(r_lt_tot, 3, total['clicks'], bold=True, bg=GREEN_LT, align='center')
cs(r_lt_tot, 4, total['impressions'], bold=True, bg=GREEN_LT, align='center')
cs(r_lt_tot, 5, total['events'], bold=True, bg=GREEN_LT, align='center')
cs(r_lt_tot, 6, f'{ctr_total}%', bold=True, bg=GREEN_LT, align='center')
ws.row_dimensions[r_lt_tot].height = 22

# Column widths
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 12

ws.sheet_view.showGridLines = False

wb.save('Phân Loại Page KNA CERT (1).xlsx')
print('Done! Sheet Báo Cáo Tháng 5 đã được tạo.')
print(f'Tổng: {total}')
