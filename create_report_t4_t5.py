import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import pathlib
src_path = pathlib.Path(__file__).parent / 'Phân Loại Page KNA CERT (1).xlsx'
wb = openpyxl.load_workbook(str(src_path))

if 'Báo Cáo T4 vs T5' in wb.sheetnames:
    del wb['Báo Cáo T4 vs T5']

ws = wb.create_sheet('Báo Cáo T4 vs T5', 0)

thin = Side(style='thin', color='BDD7EE')
thin_blue = Side(style='thin', color='4472C4')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
blue_border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)

def apply(cell, value=None, bold=False, size=10, color_font='000000', fill_color=None,
          h_align='center', v_align='center', wrap=False, border=None, num_fmt=None):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, size=size, color=color_font)
    if fill_color:
        cell.fill = PatternFill('solid', fgColor=fill_color)
    cell.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if border:
        cell.border = border
    if num_fmt:
        cell.number_format = num_fmt

def style_title(cell, text):
    apply(cell, text, bold=True, size=14, color_font='FFFFFF', fill_color='1F3864',
          h_align='center', v_align='center')

def style_section(cell, text):
    apply(cell, text, bold=True, size=11, color_font='FFFFFF', fill_color='2E75B6',
          h_align='left')

def style_header(cell, text):
    apply(cell, text, bold=True, size=10, color_font='FFFFFF', fill_color='4472C4',
          h_align='center', wrap=True, border=thin_border)

def style_data(cell, formula=None, alt=False, num_fmt=None):
    color = 'EBF3FB' if alt else 'FFFFFF'
    apply(cell, formula, fill_color=color, h_align='center', border=thin_border, num_fmt=num_fmt)

def style_label(cell, text, alt=False):
    color = 'EBF3FB' if alt else 'FFFFFF'
    apply(cell, text, bold=True, fill_color=color, h_align='left', border=thin_border)

def style_total(cell, formula=None, num_fmt=None):
    apply(cell, formula, bold=True, color_font='1F3864', fill_color='D9E1F2',
          h_align='center', border=blue_border, num_fmt=num_fmt)

def style_pct(cell, formula, alt=False):
    color = 'EBF3FB' if alt else 'FFFFFF'
    apply(cell, formula, fill_color=color, h_align='center', border=thin_border,
          num_fmt='+0.0%;-0.0%;0%')

def pct_formula(row, g_col, c_col):
    return f'=IF({c_col}{row}=0,"",({g_col}{row}-{c_col}{row})/{c_col}{row})'

def pct_total(num_c, den_c, r_start, r_end):
    return (f'=IF(SUM({den_c}{r_start}:{den_c}{r_end})=0,"",'
            f'(SUM({num_c}{r_start}:{num_c}{r_end})-SUM({den_c}{r_start}:{den_c}{r_end}))'
            f'/SUM({den_c}{r_start}:{den_c}{r_end}))')

# Column widths
col_widths = {'A':32,'B':11,'C':11,'D':13,'E':16,'F':11,'G':11,'H':13,'I':16,'J':11,'K':13,'L':16,'M':11,'N':13,'O':16}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

S4 = "'Page-Thang-4'"
S5 = "'Page Thang 5'"

nhom_list = [
    'Blog – Tin tức',
    'Service – Chứng nhận',
    'Service – Tư vấn',
    'Service – Đào tạo',
    'Trang tĩnh',
]
loai_list = [
    'Blog / Bài Viết',
    'Khách hàng',
    'Landing Page',
    'Trang Chủ',
    'Trang Danh Mục',
    'Trang Dịch Vụ',
    'Trang Liên Hệ',
    'Trang Pháp Lý',
    'Trang Về Chúng Tôi',
]

headers = [
    'Nhóm Trang',
    'Số trang T4','Click T4','Hiển Thị T4','Sự kiện QT T4',
    'Số trang T5','Click T5','Hiển Thị T5','Sự kiện QT T5',
    'Δ Click','Δ Hiển Thị','Δ Sự kiện QT',
    '% Click','% Hiển Thị','% Sự kiện QT'
]

# ============================================================
# ROW 1: TITLE
# ============================================================
ws.merge_cells('A1:O1')
style_title(ws['A1'], 'BÁO CÁO SO SÁNH THÁNG 4 vs THÁNG 5 – KNA CERT')
ws.row_dimensions[1].height = 36

# ============================================================
# NHOM TRANG
# ============================================================
r = 3
ws.merge_cells(f'A{r}:O{r}')
style_section(ws[f'A{r}'], '  I. PHÂN TÍCH THEO NHÓM TRANG')
ws.row_dimensions[r].height = 22

r = 4
for c, h in enumerate(headers, 1):
    style_header(ws.cell(row=r, column=c), h)
ws.row_dimensions[r].height = 34

data_start_nhom = r + 1
for i, nhom in enumerate(nhom_list):
    r += 1
    alt = i % 2 == 1
    style_label(ws.cell(row=r, column=1), nhom, alt)

    # T4 — Nhóm Trang = col F (6)
    style_data(ws.cell(row=r, column=2),  f'=COUNTIF({S4}!$F:$F,A{r})', alt)
    style_data(ws.cell(row=r, column=3),  f'=SUMIF({S4}!$F:$F,A{r},{S4}!$L:$L)', alt, '#,##0')
    style_data(ws.cell(row=r, column=4),  f'=SUMIF({S4}!$F:$F,A{r},{S4}!$M:$M)', alt, '#,##0')
    style_data(ws.cell(row=r, column=5),  f'=SUMIF({S4}!$F:$F,A{r},{S4}!$N:$N)', alt, '#,##0')

    # T5
    style_data(ws.cell(row=r, column=6),  f'=COUNTIF({S5}!$F:$F,A{r})', alt)
    style_data(ws.cell(row=r, column=7),  f'=SUMIF({S5}!$F:$F,A{r},{S5}!$L:$L)', alt, '#,##0')
    style_data(ws.cell(row=r, column=8),  f'=SUMIF({S5}!$F:$F,A{r},{S5}!$M:$M)', alt, '#,##0')
    style_data(ws.cell(row=r, column=9),  f'=SUMIF({S5}!$F:$F,A{r},{S5}!$N:$N)', alt, '#,##0')

    # Delta
    style_data(ws.cell(row=r, column=10), f'=G{r}-C{r}', alt, '+#,##0;-#,##0;0')
    style_data(ws.cell(row=r, column=11), f'=H{r}-D{r}', alt, '+#,##0;-#,##0;0')
    style_data(ws.cell(row=r, column=12), f'=I{r}-E{r}', alt, '+#,##0;-#,##0;0')

    # %
    style_pct(ws.cell(row=r, column=13), pct_formula(r,'G','C'), alt)
    style_pct(ws.cell(row=r, column=14), pct_formula(r,'H','D'), alt)
    style_pct(ws.cell(row=r, column=15), pct_formula(r,'I','E'), alt)
    ws.row_dimensions[r].height = 20

data_end_nhom = r

# Total row nhom
r += 1
style_total(ws.cell(row=r, column=1), 'TỔNG CỘNG')
ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
for c in range(2, 10):
    col_l = get_column_letter(c)
    style_total(ws.cell(row=r, column=c), f'=SUM({col_l}{data_start_nhom}:{col_l}{data_end_nhom})',
                '#,##0' if c >= 3 else None)
for c in range(10, 13):
    col_l = get_column_letter(c)
    style_total(ws.cell(row=r, column=c), f'=SUM({col_l}{data_start_nhom}:{col_l}{data_end_nhom})',
                '+#,##0;-#,##0;0')
# % tổng
for c, (n, d) in zip([13,14,15],[('G','C'),('H','D'),('I','E')]):
    cell = ws.cell(row=r, column=c)
    style_pct(cell, pct_total(n, d, data_start_nhom, data_end_nhom))
    cell.font = Font(bold=True)
ws.row_dimensions[r].height = 22

# ============================================================
# LOAI TRANG
# ============================================================
r += 2
ws.merge_cells(f'A{r}:O{r}')
style_section(ws[f'A{r}'], '  II. PHÂN TÍCH THEO LOẠI TRANG')
ws.row_dimensions[r].height = 22

r += 1
for c, h in enumerate(headers, 1):
    style_header(ws.cell(row=r, column=c), h)
ws.cell(row=r, column=1).value = 'Loại Trang'
ws.row_dimensions[r].height = 34

data_start_loai = r + 1
for i, loai in enumerate(loai_list):
    r += 1
    alt = i % 2 == 1
    style_label(ws.cell(row=r, column=1), loai, alt)

    # T4 — Loại Trang = col E (5)
    style_data(ws.cell(row=r, column=2),  f'=COUNTIF({S4}!$E:$E,A{r})', alt)
    style_data(ws.cell(row=r, column=3),  f'=SUMIF({S4}!$E:$E,A{r},{S4}!$L:$L)', alt, '#,##0')
    style_data(ws.cell(row=r, column=4),  f'=SUMIF({S4}!$E:$E,A{r},{S4}!$M:$M)', alt, '#,##0')
    style_data(ws.cell(row=r, column=5),  f'=SUMIF({S4}!$E:$E,A{r},{S4}!$N:$N)', alt, '#,##0')

    # T5
    style_data(ws.cell(row=r, column=6),  f'=COUNTIF({S5}!$E:$E,A{r})', alt)
    style_data(ws.cell(row=r, column=7),  f'=SUMIF({S5}!$E:$E,A{r},{S5}!$L:$L)', alt, '#,##0')
    style_data(ws.cell(row=r, column=8),  f'=SUMIF({S5}!$E:$E,A{r},{S5}!$M:$M)', alt, '#,##0')
    style_data(ws.cell(row=r, column=9),  f'=SUMIF({S5}!$E:$E,A{r},{S5}!$N:$N)', alt, '#,##0')

    # Delta
    style_data(ws.cell(row=r, column=10), f'=G{r}-C{r}', alt, '+#,##0;-#,##0;0')
    style_data(ws.cell(row=r, column=11), f'=H{r}-D{r}', alt, '+#,##0;-#,##0;0')
    style_data(ws.cell(row=r, column=12), f'=I{r}-E{r}', alt, '+#,##0;-#,##0;0')

    # %
    style_pct(ws.cell(row=r, column=13), pct_formula(r,'G','C'), alt)
    style_pct(ws.cell(row=r, column=14), pct_formula(r,'H','D'), alt)
    style_pct(ws.cell(row=r, column=15), pct_formula(r,'I','E'), alt)
    ws.row_dimensions[r].height = 20

data_end_loai = r

# Total row loai
r += 1
style_total(ws.cell(row=r, column=1), 'TỔNG CỘNG')
ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
for c in range(2, 10):
    col_l = get_column_letter(c)
    style_total(ws.cell(row=r, column=c), f'=SUM({col_l}{data_start_loai}:{col_l}{data_end_loai})',
                '#,##0' if c >= 3 else None)
for c in range(10, 13):
    col_l = get_column_letter(c)
    style_total(ws.cell(row=r, column=c), f'=SUM({col_l}{data_start_loai}:{col_l}{data_end_loai})',
                '+#,##0;-#,##0;0')
for c, (n, d) in zip([13,14,15],[('G','C'),('H','D'),('I','E')]):
    cell = ws.cell(row=r, column=c)
    style_pct(cell, pct_total(n, d, data_start_loai, data_end_loai))
    cell.font = Font(bold=True)
ws.row_dimensions[r].height = 22

# Freeze panes
ws.freeze_panes = 'B5'

import pathlib
out_path = pathlib.Path(__file__).parent / 'Phân Loại Page KNA CERT (1).xlsx'
wb.save(str(out_path))
print('Done! Sheet "Báo Cáo T4 vs T5" đã được tạo.')
print(f'Nhóm Trang: {len(nhom_list)} nhóm | Loại Trang: {len(loai_list)} loại')
