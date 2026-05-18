# -*- coding: utf-8 -*-
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT = Path(__file__).with_name("template-phan-loai-tu-khoa-b2b-dich-vu-chung-nhan.xlsx")


def style_header(cells, fill, font, border):
    for cell in cells:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def add_dropdown(ws, list_ws, headers, lists, field, list_name):
    header_to_col = {h: i + 1 for i, h in enumerate(headers)}
    list_name_to_col = {name: i + 1 for i, name in enumerate(lists.keys())}
    col = header_to_col[field]
    list_col = get_column_letter(list_name_to_col[list_name])
    end_row = 1 + len(lists[list_name])
    formula = f"='Danh_muc_dropdown'!${list_col}$2:${list_col}${end_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{ws.max_row}")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Phan_loai_tu_khoa"

    headers = [
        "tu_khoa",
        "tu_khoa_chuan_hoa",
        "nhom_dich_vu",
        "tieu_chuan_chung_nhan",
        "nganh_ap_dung",
        "doi_tuong_khach_hang",
        "y_dinh_chinh",
        "y_dinh_phu",
        "giai_doan_pheu",
        "loai_trang",
        "cum_chu_de",
        "chu_de_con",
        "tu_khoa_chinh",
        "vai_tro_tu_khoa",
        "goc_noi_dung",
        "gia_dinh_serp",
        "gia_tri_kinh_doanh",
        "uu_tien_seo",
        "do_tin_cay",
        "ghi_chu",
    ]
    ws.append(headers)

    sample_rows = [
        [
            "dịch vụ chứng nhận iso 9001",
            "dịch vụ chứng nhận iso 9001",
            "Chứng nhận ISO",
            "ISO 9001",
            "Đa ngành",
            "Chủ doanh nghiệp/QA Manager",
            "Giao dịch",
            "Địa phương",
            "BOFU",
            "Trang sản phẩm/dịch vụ",
            "Chứng nhận ISO 9001",
            "Dịch vụ chứng nhận",
            "dịch vụ chứng nhận iso 9001",
            "Chính",
            "Nêu quy trình, thời gian, hồ sơ và CTA tư vấn",
            "SERP có nhà cung cấp dịch vụ và bài giải thích",
            "Cao",
            "Cao",
            "Cao",
            "Keyword lead trực tiếp",
        ],
        [
            "iso 9001 là gì",
            "iso 9001 là gì",
            "Chứng nhận ISO",
            "ISO 9001",
            "Đa ngành",
            "Người mới tìm hiểu/QA",
            "Thông tin",
            "",
            "TOFU",
            "Bài blog",
            "Chứng nhận ISO 9001",
            "Khái niệm tiêu chuẩn",
            "iso 9001 là gì",
            "Chính",
            "Giải thích tiêu chuẩn, lợi ích, đối tượng cần áp dụng",
            "SERP thiên về định nghĩa và hướng dẫn",
            "Trung bình",
            "Trung bình",
            "Cao",
            "Dùng để kéo internal link về trang dịch vụ",
        ],
        [
            "chi phí chứng nhận haccp",
            "chi phí chứng nhận haccp",
            "Chứng nhận an toàn thực phẩm",
            "HACCP",
            "Thực phẩm",
            "Nhà máy/QA/QC",
            "Thương mại",
            "Giao dịch",
            "MOFU",
            "Landing page",
            "Chứng nhận HACCP",
            "Chi phí và báo giá",
            "chi phí chứng nhận haccp",
            "Chính",
            "Trình bày yếu tố ảnh hưởng chi phí và form nhận báo giá",
            "SERP có bài giá, dịch vụ và tư vấn",
            "Cao",
            "Cao",
            "Trung bình",
            "Cần kiểm tra SERP để quyết định landing hay bài tư vấn",
        ],
        [
            "tổ chức chứng nhận iso uy tín",
            "tổ chức chứng nhận iso uy tín",
            "Chứng nhận ISO",
            "ISO",
            "Đa ngành",
            "Procurement/Chủ doanh nghiệp",
            "Thương mại",
            "",
            "MOFU",
            "Trang so sánh",
            "Lựa chọn đơn vị chứng nhận",
            "Tiêu chí chọn nhà cung cấp",
            "tổ chức chứng nhận iso uy tín",
            "Chính",
            "Tiêu chí đánh giá năng lực, công nhận, kinh nghiệm ngành",
            "SERP có list đơn vị và bài tư vấn",
            "Cao",
            "Cao",
            "Trung bình",
            "Cần tránh claim quá đà",
        ],
        [
            "chứng nhận ce cho thiết bị y tế",
            "chứng nhận ce cho thiết bị y tế",
            "Chứng nhận xuất khẩu",
            "CE",
            "Thiết bị y tế",
            "Doanh nghiệp xuất khẩu/Regulatory",
            "Thương mại",
            "Giao dịch",
            "MOFU",
            "Trang ngành",
            "Chứng nhận CE",
            "Thiết bị y tế",
            "chứng nhận ce cho thiết bị y tế",
            "Chính",
            "Giải thích yêu cầu, hồ sơ, quy trình và tư vấn chuyên ngành",
            "SERP có hướng dẫn pháp lý và nhà cung cấp",
            "Cao",
            "Cao",
            "Trung bình",
            "Cần xác minh yêu cầu pháp lý theo thị trường",
        ],
    ]
    for row in sample_rows:
        ws.append(row)
    for _ in range(250):
        ws.append([""] * len(headers))

    header_fill = PatternFill("solid", fgColor="1F4E79")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    style_header(ws[1], header_fill, white_font, border)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    widths = {
        1: 30,
        2: 30,
        3: 26,
        4: 22,
        5: 24,
        6: 28,
        7: 18,
        8: 18,
        9: 16,
        10: 24,
        11: 28,
        12: 26,
        13: 30,
        14: 18,
        15: 42,
        16: 42,
        17: 18,
        18: 16,
        19: 14,
        20: 42,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[1].height = 34

    tab = Table(displayName="BangPhanLoaiTuKhoa", ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    comments = {
        "tu_khoa": "Từ khóa gốc từ Ahrefs, SEMrush, GSC, Keyword Planner hoặc brainstorm.",
        "nhom_dich_vu": "Nhóm dịch vụ lớn như Chứng nhận ISO, chứng nhận hợp quy, kiểm định, audit, chứng nhận xuất khẩu.",
        "tieu_chuan_chung_nhan": "Ví dụ: ISO 9001, ISO 14001, ISO 22000, HACCP, GMP, CE, FDA, FSC, BSCI, WRAP.",
        "y_dinh_chinh": "Intent chính của người tìm kiếm.",
        "loai_trang": "Loại page nên dùng để target keyword này.",
        "tu_khoa_chinh": "Keyword đại diện cho trang. Các biến thể cùng intent sẽ map về keyword chính này.",
        "gia_tri_kinh_doanh": "Đánh giá mức liên quan đến lead/doanh thu.",
        "uu_tien_seo": "Ưu tiên triển khai sau khi cân nhắc intent, business value và khả năng ranking.",
        "ghi_chu": "Ghi rủi ro cannibalization, cần check SERP, cần xác minh pháp lý hoặc bối cảnh kinh doanh.",
    }
    for col_idx, header in enumerate(headers, 1):
        if header in comments:
            ws.cell(row=1, column=col_idx).comment = Comment(comments[header], "Codex")

    list_ws = wb.create_sheet("Danh_muc_dropdown")
    lists = {
        "nhom_dich_vu": ["Chứng nhận ISO", "Chứng nhận an toàn thực phẩm", "Chứng nhận hợp quy", "Kiểm định", "Audit/đánh giá", "Chứng nhận xuất khẩu", "Tư vấn hệ thống", "Đào tạo", "Khác"],
        "tieu_chuan_chung_nhan": ["ISO", "ISO 9001", "ISO 14001", "ISO 22000", "ISO 45001", "ISO 13485", "HACCP", "GMP", "CE", "FDA", "FSC", "BSCI", "WRAP", "Hợp quy", "Khác"],
        "nganh_ap_dung": ["Đa ngành", "Thực phẩm", "Dệt may", "Cơ khí", "Thiết bị y tế", "Mỹ phẩm", "Nông sản", "Bao bì", "Điện/điện tử", "Xây dựng", "Xuất khẩu", "Khác"],
        "doi_tuong_khach_hang": ["Chủ doanh nghiệp", "QA Manager", "QA/QC", "Procurement", "Regulatory", "Nhà máy", "SME", "Doanh nghiệp xuất khẩu", "Phòng vận hành", "Khác"],
        "y_dinh": ["Thông tin", "Thương mại", "Giao dịch", "Điều hướng", "Địa phương", "Hỗ trợ"],
        "giai_doan_pheu": ["TOFU", "MOFU", "BOFU", "Retention"],
        "loai_trang": ["Bài blog", "Trang pillar", "Trang danh mục", "Trang sản phẩm/dịch vụ", "Landing page", "Trang so sánh", "FAQ/hỗ trợ", "Trang địa phương", "Trang ngành", "Case study", "Trang chủ/brand"],
        "vai_tro_tu_khoa": ["Chính", "Phụ", "Supporting", "FAQ", "Semantic", "Biến thể địa phương", "Biến thể ngành"],
        "muc_do": ["Cao", "Trung bình", "Thấp"],
        "do_tin_cay": ["Cao", "Trung bình", "Thấp"],
    }
    for col_idx, (name, values) in enumerate(lists.items(), 1):
        cell = list_ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = white_font
        list_ws.column_dimensions[get_column_letter(col_idx)].width = max(22, max(len(v) for v in values) + 2)
        for row_idx, value in enumerate(values, 2):
            list_ws.cell(row=row_idx, column=col_idx, value=value)
    list_ws.freeze_panes = "A2"

    for field, list_name in [
        ("nhom_dich_vu", "nhom_dich_vu"),
        ("tieu_chuan_chung_nhan", "tieu_chuan_chung_nhan"),
        ("nganh_ap_dung", "nganh_ap_dung"),
        ("doi_tuong_khach_hang", "doi_tuong_khach_hang"),
        ("y_dinh_chinh", "y_dinh"),
        ("y_dinh_phu", "y_dinh"),
        ("giai_doan_pheu", "giai_doan_pheu"),
        ("loai_trang", "loai_trang"),
        ("vai_tro_tu_khoa", "vai_tro_tu_khoa"),
        ("gia_tri_kinh_doanh", "muc_do"),
        ("uu_tien_seo", "muc_do"),
        ("do_tin_cay", "do_tin_cay"),
    ]:
        add_dropdown(ws, list_ws, headers, lists, field, list_name)

    guide = wb.create_sheet("Huong_dan")
    guide_rows = [
        ["Template phân loại từ khóa SEO - B2B dịch vụ chứng nhận"],
        ["Mục tiêu", "Map từ khóa về đúng intent, đúng loại trang và đúng cụm chủ đề để lập kế hoạch SEO có thể triển khai."],
        ["Bước 1", "Dán danh sách keyword vào cột tu_khoa."],
        ["Bước 2", "Chuẩn hóa keyword, gán nhóm dịch vụ, tiêu chuẩn chứng nhận, ngành áp dụng và đối tượng khách hàng."],
        ["Bước 3", "Phân loại intent, funnel stage và loại trang phù hợp."],
        ["Bước 4", "Gộp keyword cùng intent vào cùng tu_khoa_chinh; tách page khi intent hoặc ngành áp dụng khác nhau."],
        ["Bước 5", "Chấm business value và SEO priority. Ưu tiên keyword có khả năng tạo lead, dịch vụ rõ và intent BOFU/MOFU."],
        ["Lưu ý", "Các keyword pháp lý/chứng nhận nên ghi chú khi cần xác minh quy định, phạm vi công nhận hoặc thị trường xuất khẩu."],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.merge_cells("A1:B1")
    guide["A1"].fill = header_fill
    guide["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 110

    ex = wb.create_sheet("Vi_du")
    ex.append(headers)
    for row in sample_rows:
        ex.append(row)
    style_header(ex[1], header_fill, white_font, border)
    for row in ex.iter_rows(min_row=1, max_row=ex.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, width in widths.items():
        ex.column_dimensions[get_column_letter(idx)].width = width
    ex.freeze_panes = "A2"

    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, read_only=True)
    assert check["Phan_loai_tu_khoa"]["A2"].value == "dịch vụ chứng nhận iso 9001"
    assert check["Huong_dan"]["A1"].value == "Template phân loại từ khóa SEO - B2B dịch vụ chứng nhận"
    print(OUTPUT)


if __name__ == "__main__":
    main()
