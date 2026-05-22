import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, pathlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = pathlib.Path(__file__).parent / 'Phân Loại Page KNA CERT (1).xlsx'
wb = openpyxl.load_workbook(str(src))
ws = wb['Báo Cáo Tháng 5']

# Đọc dữ liệu thực từ Page Thang 5 để tính nhận xét
ws5 = wb['Page Thang 5']

def aggregate(ws, group_col, click_col=12, ht_col=13, sq_col=14):
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[group_col - 1]
        if not key:
            continue
        key = str(key).strip()
        c = row[click_col - 1] or 0
        h = row[ht_col - 1] or 0
        s = row[sq_col - 1] or 0
        if key not in result:
            result[key] = {'pages': 0, 'click': 0, 'hien_thi': 0, 'sq': 0}
        result[key]['pages'] += 1
        result[key]['click'] += c
        result[key]['hien_thi'] += h
        result[key]['sq'] += s
    return result

nhom5 = aggregate(ws5, 6)
loai5 = aggregate(ws5, 5)
tong_click = sum(v['click'] for v in nhom5.values())
tong_ht    = sum(v['hien_thi'] for v in nhom5.values())
tong_sq    = sum(v['sq'] for v in nhom5.values())
tong_ctr   = tong_click / tong_ht * 100 if tong_ht else 0

# Tìm dòng cuối sheet
last_row = ws.max_row
start_r  = last_row + 2

thin      = Side(style='thin', color='BDD7EE')
thin_blue = Side(style='thin', color='4472C4')
b_thin    = Border(left=thin,      right=thin,      top=thin,      bottom=thin)
b_blue    = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)

MAX_COL = 6   # sheet chỉ có cột A–F

def merge(r, text, bold=True, size=11, font_color='FFFFFF', bg='1F3864', indent=1, h=22):
    ws.merge_cells(f'A{r}:F{r}')
    c = ws[f'A{r}']
    c.value = text
    c.font = Font(bold=bold, size=size, color=font_color)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=indent)
    ws.row_dimensions[r].height = h

def bullet(r, label, text, good=None, alt=False):
    # A–B: label
    ws.merge_cells(f'A{r}:B{r}')
    lc = ws[f'A{r}']
    lc.value = label
    lc.font = Font(bold=True, size=10,
                   color=('9C0006' if good is False else ('375623' if good is True else '1F3864')))
    lc.fill = PatternFill('solid', fgColor='EBF3FB' if alt else 'F2F7FB')
    lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    lc.border = b_thin

    # C–F: nội dung
    ws.merge_cells(f'C{r}:F{r}')
    tc = ws[f'C{r}']
    tc.value = text
    if good is True:
        tc.font = Font(size=10, color='375623')
        tc.fill = PatternFill('solid', fgColor='E2EFDA')
    elif good is False:
        tc.font = Font(size=10, color='9C0006')
        tc.fill = PatternFill('solid', fgColor='FFCCCC')
    else:
        tc.font = Font(size=10, color='1F3864')
        tc.fill = PatternFill('solid', fgColor='EBF3FB' if alt else 'F2F7FB')
    tc.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    tc.border = b_thin
    ws.row_dimensions[r].height = 22

# ============================================================
r = start_r
merge(r, '  NHẬN XÉT THÁNG 5 – KNA CERT', size=12, bg='1F3864', h=28)

# ----- A. TỔNG QUAN -----
r += 1
merge(r, '  A. TỔNG QUAN', size=11, bg='2E75B6')

r += 1
bullet(r, '  Tổng lưu lượng',
    f'726 trang được phân tích. Tổng click: {tong_click:,} | Tổng hiển thị: {tong_ht:,} | '
    f'CTR trung bình: {tong_ctr:.2f}%. CTR ở mức thấp – còn nhiều dư địa cải thiện tiêu đề và mô tả.',
    good=None)

r += 1
# Nhóm CTR cao nhất
best_ctr_nhom = max(nhom5.items(), key=lambda x: x[1]['click']/x[1]['hien_thi'] if x[1]['hien_thi'] else 0)
k, v = best_ctr_nhom
ctr_val = v['click']/v['hien_thi']*100 if v['hien_thi'] else 0
bullet(r, '  Nhóm CTR cao nhất',
    f'{k}: CTR {ctr_val:.2f}% ({v["click"]:,} click / {v["hien_thi"]:,} HT). '
    f'Đây là nhóm hiệu quả nhất về khả năng thu hút click từ kết quả tìm kiếm.',
    good=True)

r += 1
# Nhóm SQ nhiều nhất
best_sq = max(nhom5.items(), key=lambda x: x[1]['sq'])
k2, v2 = best_sq
bullet(r, '  Sự kiện QT cao nhất',
    f'Nhóm {k2} dẫn đầu với {v2["sq"]} sự kiện quan trọng – '
    f'cho thấy nhóm này đang tạo ra hành động thực tế từ người dùng nhiều nhất.',
    good=True)

# ----- B. NHÓM TRANG -----
r += 1
merge(r, '  B. NHÓM TRANG', size=11, bg='2E75B6')

nhom_order = ['Blog – Tin tức', 'Service – Chứng nhận', 'Service – Tư vấn', 'Service – Đào tạo', 'Trang tĩnh']
comments_nhom = {
    'Blog – Tin tức': (
        True,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Nhóm lớn nhất, chiếm {v["hien_thi"]/tong_ht*100:.1f}% tổng Hiển Thị. '
            f'CTR 1.3% còn thấp so với tiềm năng – nên tối ưu title/meta description cho các bài viết có HT cao nhưng ít click.'
        )
    ),
    'Service – Chứng nhận': (
        False,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'CTR thấp nhất trong tất cả nhóm (0.80%) và sự kiện QT chỉ đạt {v["sq"]}. '
            f'Ưu tiên rà soát từ khóa, cập nhật nội dung và cải thiện CTA trên các trang chứng nhận.'
        )
    ),
    'Service – Tư vấn': (
        None,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Nhóm nhỏ nhất (13 trang) nhưng CTR đạt 1.03%. '
            f'Hiển Thị còn hạn chế – cần đẩy mạnh nội dung và liên kết nội bộ để tăng độ phủ từ khóa tư vấn.'
        )
    ),
    'Service – Đào tạo': (
        True,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Nhóm hiệu quả nhất: CTR cao nhất {v["click"]/v["hien_thi"]*100:.2f}% và SQ dẫn đầu {v["sq"]} sự kiện. '
            f'Nên nhân rộng mô hình nội dung của nhóm này sang các nhóm service khác.'
        )
    ),
    'Trang tĩnh': (
        None,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Lưu lượng thấp do đặc thù nhóm (trang pháp lý, liên hệ, về chúng tôi...). '
            f'Tập trung tối ưu UX hơn là SEO cho nhóm này.'
        )
    ),
}

for i, nhom in enumerate(nhom_order):
    r += 1
    v = nhom5.get(nhom, {'pages':0,'click':0,'hien_thi':0,'sq':0})
    good, fn = comments_nhom[nhom]
    bullet(r, f'  {nhom}', fn(v), good=good, alt=(i % 2 == 1))

# ----- C. LOẠI TRANG -----
r += 1
merge(r, '  C. LOẠI TRANG', size=11, bg='2E75B6')

loai_order = [
    'Blog / Bài Viết', 'Trang Dịch Vụ', 'Landing Page',
    'Trang Danh Mục', 'Khách hàng', 'Trang Về Chúng Tôi',
    'Trang Liên Hệ', 'Trang Pháp Lý', 'Trang Chủ',
]
comments_loai = {
    'Blog / Bài Viết': (
        True,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Chiếm {v["hien_thi"]/tong_ht*100:.1f}% tổng Hiển Thị toàn site. '
            f'CTR 1.30% còn thấp – ưu tiên tối ưu title/meta cho top 50 bài có HT cao nhưng CTR thấp.'
        )
    ),
    'Trang Dịch Vụ': (
        False,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Chiếm {v["hien_thi"]/tong_ht*100:.1f}% tổng HT nhưng SQ chỉ đạt {v["sq"]}. '
            f'CTR 1.65% và SQ thấp báo hiệu trang dịch vụ chưa đủ thuyết phục – cần bổ sung FAQ, testimonial và CTA rõ ràng hơn.'
        )
    ),
    'Landing Page': (
        True,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'CTR cao nhất toàn site ({v["click"]/v["hien_thi"]*100:.2f}%) và SQ dẫn đầu ({v["sq"]} sự kiện). '
            f'Đây là loại trang hiệu quả nhất – nên tăng số lượng landing page cho các dịch vụ chủ lực.'
        )
    ),
    'Trang Danh Mục': (
        None,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'CTR 3.59% tốt nhưng HT chỉ {v["hien_thi"]:,} – độ phủ từ khóa rất hạn chế. '
            f'Bổ sung nội dung mô tả danh mục và internal link để tăng HT.'
        )
    ),
    'Khách hàng': (
        False,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'Lưu lượng rất thấp dù có {v["pages"]} trang. '
            f'Cần cập nhật case study thực tế, thêm schema Review và tối ưu từ khóa thương hiệu khách hàng.'
        )
    ),
    'Trang Về Chúng Tôi': (
        None,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | CTR: {v["click"]/v["hien_thi"]*100:.2f}% | SQ: {v["sq"]}. '
            f'CTR 1.44% ổn. Cần tăng E-E-A-T: thêm thông tin đội ngũ, chứng chỉ, lịch sử hoạt động để cải thiện HT.'
        )
    ),
    'Trang Liên Hệ': (
        True,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | SQ: {v["sq"]}. '
            f'Sự kiện QT đạt {v["sq"]} – cho thấy traffic vào trang liên hệ đang chuyển đổi tốt. '
            f'Đảm bảo form liên hệ và số điện thoại luôn hoạt động.'
        )
    ),
    'Trang Pháp Lý': (
        None,
        lambda v: (
            f'{v["pages"]} trang | HT: {v["hien_thi"]:,} – lưu lượng không đáng kể, bình thường với loại trang này.'
        )
    ),
    'Trang Chủ': (
        False,
        lambda v: (
            f'{v["pages"]} trang | Click: {v["click"]:,} | HT: {v["hien_thi"]:,} | SQ: {v["sq"]}. '
            f'Click và Hiển Thị = 0 là bất thường. Kiểm tra lại dữ liệu Google Search Console – '
            f'trang chủ thường có lưu lượng từ khóa thương hiệu.'
        )
    ),
}

for i, loai in enumerate(loai_order):
    r += 1
    v = loai5.get(loai, {'pages':0,'click':0,'hien_thi':0,'sq':0})
    good, fn = comments_loai[loai]
    bullet(r, f'  {loai}', fn(v), good=good, alt=(i % 2 == 1))

# ----- D. ƯU TIÊN -----
r += 1
merge(r, '  D. ƯU TIÊN HÀNH ĐỘNG', size=11, bg='375623', h=22)

actions = [
    (False, '  [Ưu tiên 1]',
     'Tối ưu CTR Blog/Bài Viết: rà soát top 30 bài có HT > 500 nhưng CTR < 1% → viết lại title và meta description hấp dẫn hơn.'),
    (False, '  [Ưu tiên 2]',
     'Phục hồi Service – Chứng nhận: kiểm tra từ khóa mất thứ hạng, cập nhật nội dung, bổ sung schema và backlink chất lượng.'),
    (True,  '  [Ưu tiên 3]',
     'Nhân rộng Landing Page: Service – Đào tạo đang có CTR và SQ tốt nhất – tạo thêm landing page cho các khóa đào tạo mới và chứng nhận chủ lực.'),
    (True,  '  [Ưu tiên 4]',
     'Kiểm tra Trang Chủ: HT = 0 cần xác minh lại filter GSC – nếu đúng thì cần tối ưu từ khóa thương hiệu và homepage SEO ngay.'),
]
for i, (good, label, text) in enumerate(actions):
    r += 1
    bullet(r, label, text, good=good, alt=(i % 2 == 1))

ws.row_dimensions[r].height = 22

wb.save(str(src))
print(f'Done! Đã thêm nhận xét từ row {start_r} đến row {r}.')
