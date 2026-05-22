import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, pathlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = pathlib.Path(__file__).parent / 'Phân Loại Page KNA CERT (1).xlsx'
wb = openpyxl.load_workbook(str(src))
ws = wb['Báo Cáo Tháng 5']

S5 = "'Page Thang 5'"

# Cột trong Page Thang 5:
# C = URL, E = Loại Trang, F = Nhóm Trang
# L = Click, M = Hiển Thị, N = Sự kiện QT

PCT_FMT = '0.00%'

def set_formula(cell, formula, num_fmt=None):
    cell.value = formula
    if num_fmt:
        cell.number_format = num_fmt

# ============================================================
# TỔNG TOÀN WEBSITE — row 6
# ============================================================
# B6: Số trang (đếm URL không rỗng, trừ header)
set_formula(ws['B6'], f'=COUNTA({S5}!$C:$C)-1')
# C6: Tổng Click
set_formula(ws['C6'], f'=SUM({S5}!$L:$L)', '#,##0')
# D6: Tổng Hiển Thị
set_formula(ws['D6'], f'=SUM({S5}!$M:$M)', '#,##0')
# E6: Tổng Sự kiện
set_formula(ws['E6'], f'=SUM({S5}!$N:$N)', '#,##0')
# F6: CTR = Click / Hiển Thị  (cột F, trước đây là E)
# Kiểm tra sheet có cột F không — nếu không thì dùng E
# Theo data hiện tại: A5=Số trang, B5=Click, C5=HT, D5=SQ, E5=CTR
set_formula(ws['E6'], f'=SUM({S5}!$N:$N)', '#,##0')

# CTR ở cột E6 hiện là '1.49%' (text) — ghi đè bằng công thức
# Cấu trúc: A=Số trang, B=Click, C=HT, D=SQ, E=CTR
ws['A6'].value = f'=COUNTA({S5}!$C:$C)-1'
ws['B6'].value = f'=SUM({S5}!$L:$L)'
ws['C6'].value = f'=SUM({S5}!$M:$M)'
ws['D6'].value = f'=SUM({S5}!$N:$N)'
ws['E6'].value = f'=IF(C6=0,0,B6/C6)'
ws['E6'].number_format = PCT_FMT
for col in ['A','B','C','D']:
    ws[f'{col}6'].number_format = '#,##0'

# ============================================================
# NHÓM TRANG — rows 10–14, tổng row 15
# ============================================================
nhom_rows = {
    10: 'Blog – Tin tức',
    11: 'Service – Chứng nhận',
    12: 'Service – Tư vấn',
    13: 'Service – Đào tạo',
    14: 'Trang tĩnh',
}

for r, nhom in nhom_rows.items():
    # A = tên nhóm (giữ nguyên text)
    # B = Số trang
    ws[f'B{r}'].value = f'=COUNTIF({S5}!$F:$F,A{r})'
    ws[f'B{r}'].number_format = '#,##0'
    # C = Click
    ws[f'C{r}'].value = f'=SUMIF({S5}!$F:$F,A{r},{S5}!$L:$L)'
    ws[f'C{r}'].number_format = '#,##0'
    # D = Hiển Thị
    ws[f'D{r}'].value = f'=SUMIF({S5}!$F:$F,A{r},{S5}!$M:$M)'
    ws[f'D{r}'].number_format = '#,##0'
    # E = Sự kiện
    ws[f'E{r}'].value = f'=SUMIF({S5}!$F:$F,A{r},{S5}!$N:$N)'
    ws[f'E{r}'].number_format = '#,##0'
    # F = CTR
    ws[f'F{r}'].value = f'=IF(D{r}=0,0,C{r}/D{r})'
    ws[f'F{r}'].number_format = PCT_FMT

# Tổng nhóm row 15
ws['B15'].value = '=SUM(B10:B14)'
ws['C15'].value = '=SUM(C10:C14)'
ws['D15'].value = '=SUM(D10:D14)'
ws['E15'].value = '=SUM(E10:E14)'
ws['F15'].value = '=IF(D15=0,0,C15/D15)'
ws['F15'].number_format = PCT_FMT
for col in ['B','C','D','E']:
    ws[f'{col}15'].number_format = '#,##0'

# ============================================================
# LOẠI TRANG — rows 19–27, tổng row 28
# ============================================================
loai_rows = {
    19: 'Blog / Bài Viết',
    20: 'Khách hàng',
    21: 'Landing Page',
    22: 'Trang Chủ',
    23: 'Trang Danh Mục',
    24: 'Trang Dịch Vụ',
    25: 'Trang Liên Hệ',
    26: 'Trang Pháp Lý',
    27: 'Trang Về Chúng Tôi',
}

for r, loai in loai_rows.items():
    ws[f'B{r}'].value = f'=COUNTIF({S5}!$E:$E,A{r})'
    ws[f'B{r}'].number_format = '#,##0'
    ws[f'C{r}'].value = f'=SUMIF({S5}!$E:$E,A{r},{S5}!$L:$L)'
    ws[f'C{r}'].number_format = '#,##0'
    ws[f'D{r}'].value = f'=SUMIF({S5}!$E:$E,A{r},{S5}!$M:$M)'
    ws[f'D{r}'].number_format = '#,##0'
    ws[f'E{r}'].value = f'=SUMIF({S5}!$E:$E,A{r},{S5}!$N:$N)'
    ws[f'E{r}'].number_format = '#,##0'
    ws[f'F{r}'].value = f'=IF(D{r}=0,0,C{r}/D{r})'
    ws[f'F{r}'].number_format = PCT_FMT

# Tổng loại row 28
ws['B28'].value = '=SUM(B19:B27)'
ws['C28'].value = '=SUM(C19:C27)'
ws['D28'].value = '=SUM(D19:D27)'
ws['E28'].value = '=SUM(E19:E27)'
ws['F28'].value = '=IF(D28=0,0,C28/D28)'
ws['F28'].number_format = PCT_FMT
for col in ['B','C','D','E']:
    ws[f'{col}28'].number_format = '#,##0'

wb.save(str(src))
print('Done! Sheet "Báo Cáo Tháng 5" đã được liên kết với "Page Thang 5" bằng công thức.')
