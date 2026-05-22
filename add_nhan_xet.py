import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl, pathlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

src = pathlib.Path(__file__).parent / 'Phân Loại Page KNA CERT (1).xlsx'
wb = openpyxl.load_workbook(str(src))
ws = wb['Báo Cáo T4 vs T5']

# Tìm dòng cuối cùng có dữ liệu
last_row = ws.max_row
start_r = last_row + 2  # cách 2 dòng

thin = Side(style='thin', color='BDD7EE')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def write_section_title(row, text, color='1F3864'):
    ws.merge_cells(f'A{row}:O{row}')
    cell = ws[f'A{row}']
    cell.value = text
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=color)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 24

def write_sub_title(row, text):
    ws.merge_cells(f'A{row}:O{row}')
    cell = ws[f'A{row}']
    cell.value = text
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='2E75B6')
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22

def write_bullet(row, label, text, label_color='1F3864', good=None):
    # Cột A: icon + label (merge A:B)
    ws.merge_cells(f'A{row}:B{row}')
    lc = ws[f'A{row}']
    lc.value = label
    lc.font = Font(bold=True, size=10, color=label_color)
    lc.fill = PatternFill('solid', fgColor='F2F7FB')
    lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    lc.border = border

    # Cột C–O: nội dung
    ws.merge_cells(f'C{row}:O{row}')
    tc = ws[f'C{row}']
    tc.value = text
    if good is True:
        tc.font = Font(size=10, color='375623')
        tc.fill = PatternFill('solid', fgColor='E2EFDA')
    elif good is False:
        tc.font = Font(size=10, color='9C0006')
        tc.fill = PatternFill('solid', fgColor='FFCCCC')
    else:
        tc.font = Font(size=10, color='1F3864')
        tc.fill = PatternFill('solid', fgColor='F2F7FB')
    tc.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    tc.border = border
    ws.row_dimensions[row].height = 22

def write_empty(row):
    ws.row_dimensions[row].height = 6

# ============================================================
r = start_r
write_section_title(r, '  NHẬN XÉT & ĐÁNH GIÁ THÁNG 4 vs THÁNG 5', '1F3864')

# ------- TỔNG QUAN -------
r += 1
write_sub_title(r, '  A. TỔNG QUAN TOÀN WEBSITE')

r += 1
write_bullet(r, '  Tổng Click',
    'Tháng 5 đạt 2.298 click, giảm nhẹ -2,5% so với tháng 4 (2.357 click). '
    'Mức giảm không đáng kể, nằm trong dao động bình thường theo mùa vụ.', good=None)

r += 1
write_bullet(r, '  Tổng Hiển Thị',
    'Hiển thị gần như giữ nguyên: T4 = 153.590 → T5 = 153.963 (+0,2%). '
    'Cho thấy độ phủ từ khóa của site ổn định, không bị mất thứ hạng đáng kể.', good=True)

r += 1
write_bullet(r, '  Sự Kiện QT',
    'Sự kiện quan trọng tăng từ 29 → 32 (+10,3%). '
    'Tín hiệu tích cực về chất lượng traffic dù lượng click giảm nhẹ.', good=True)

write_empty(r + 1)

# ------- NHÓM TRANG -------
r += 2
write_sub_title(r, '  B. THEO NHÓM TRANG')

r += 1
write_bullet(r, '  Blog – Tin tức',
    '359 trang | Click tăng +0,5% (1.087 → 1.092), Hiển Thị tăng +6,7% (78.872 → 84.194). '
    'Sự kiện QT tăng mạnh +80% (5 → 9). Nhóm đang cải thiện đều – cần duy trì tần suất đăng bài.',
    good=True)

r += 1
write_bullet(r, '  Service – Chứng nhận',
    '114 trang | Click giảm -28,5% (284 → 203), Hiển Thị giảm -16,0% (30.048 → 25.231). '
    'Sự kiện QT giảm rất mạnh -84,6% (13 → 2). Đây là nhóm đáng lo ngại nhất – cần kiểm tra lại thứ hạng từ khóa và nội dung trang dịch vụ chứng nhận.',
    label_color='9C0006', good=False)

r += 1
write_bullet(r, '  Service – Tư vấn',
    '13 trang | Click tăng +20,6% (34 → 41) dù Hiển Thị giảm -21,7% (5.074 → 3.973). '
    'CTR cải thiện rõ rệt – tiêu đề/mô tả đang hút click tốt hơn mặc dù ít xuất hiện hơn.',
    good=True)

r += 1
write_bullet(r, '  Service – Đào tạo',
    '117 trang | Click +0,7% (913 → 919), Hiển Thị +7,5% (34.567 → 37.143). '
    'Sự kiện QT tăng +45,5% (11 → 16). Nhóm ổn định và có xu hướng tốt lên – tiếp tục đẩy SEO các trang khóa đào tạo.',
    good=True)

r += 1
write_bullet(r, '  Trang tĩnh',
    '123 trang | Click tăng +10,3% (39 → 43) nhưng Hiển Thị giảm -32,0% (5.029 → 3.422). '
    'CTR tăng mạnh, tuy nhiên Hiển Thị giảm cần theo dõi thêm – có thể do thay đổi thuật toán hoặc cạnh tranh từ khóa.',
    good=None)

write_empty(r + 1)

# ------- LOẠI TRANG -------
r += 2
write_sub_title(r, '  C. THEO LOẠI TRANG')

r += 1
write_bullet(r, '  Blog / Bài Viết',
    '348 trang – nhóm lớn nhất. Hiển Thị tăng +6,8% (78.791 → 84.154), Sự kiện QT +80%. '
    'Bài viết đang được Google tín nhiệm hơn, nên duy trì chất lượng nội dung và tối ưu internal link.',
    good=True)

r += 1
write_bullet(r, '  Trang Dịch Vụ',
    '181 trang – nhóm lớn thứ 2. Click giảm -6,9% (1.138 → 1.060), Hiển Thị giảm -5,0% (67.445 → 64.079). '
    'Sự kiện QT giảm mạnh -75% (16 → 4). Cần rà soát các trang dịch vụ chủ lực, kiểm tra title/description và CTA.',
    label_color='9C0006', good=False)

r += 1
write_bullet(r, '  Landing Page',
    '45 trang | Click +9,8% (92 → 101), Hiển Thị +2,0%. Sự kiện QT tăng +75% (8 → 14). '
    'Trang landing đang chuyển đổi hiệu quả hơn – đây là tín hiệu rất tích cực cho mục tiêu kinh doanh.',
    good=True)

r += 1
write_bullet(r, '  Trang Danh Mục',
    '30 trang | Click tăng +100% (3 → 6) nhưng Hiển Thị giảm -35% (257 → 167). '
    'Lượng hiển thị giảm nhưng những lần xuất hiện lại hiệu quả hơn. Cần bổ sung nội dung và internal link cho các trang danh mục.',
    good=None)

r += 1
write_bullet(r, '  Trang Về Chúng Tôi',
    '15 trang | Click tăng +40% (15 → 21) nhưng Hiển Thị giảm -32,4% (2.157 → 1.458). '
    'CTR tăng mạnh cho thấy nội dung đang phù hợp hơn với intent người dùng. '
    'Tuy nhiên Hiển Thị giảm cần tối ưu thêm từ khóa thương hiệu và E-E-A-T.',
    good=None)

r += 1
write_bullet(r, '  Khách Hàng',
    '95 trang | Click giảm -19,0% (21 → 17), Hiển Thị giảm -31,9% (2.582 → 1.758). '
    'Trang khách hàng/case study đang mất dần lưu lượng – cần cập nhật nội dung và thêm schema review.',
    label_color='9C0006', good=False)

write_empty(r + 1)

# ------- ƯU TIÊN HÀNH ĐỘNG -------
r += 2
write_section_title(r, '  D. ƯU TIÊN HÀNH ĐỘNG THÁNG 6', '375623')

r += 1
write_bullet(r, '  [Ưu tiên 1]',
    'Kiểm tra & phục hồi nhóm Service – Chứng nhận: rà soát từ khóa mất thứ hạng, '
    'cập nhật nội dung trang dịch vụ, tối ưu on-page và xây dựng thêm backlink.', label_color='9C0006', good=False)

r += 1
write_bullet(r, '  [Ưu tiên 2]',
    'Cải thiện Trang Dịch Vụ: review title/meta description, bổ sung FAQ schema, '
    'kiểm tra CTA và tối ưu UX để phục hồi Sự kiện QT (đã giảm -75%).', label_color='9C0006', good=False)

r += 1
write_bullet(r, '  [Ưu tiên 3]',
    'Phát huy Blog – Tin tức: duy trì lịch đăng bài đều đặn, tập trung topical authority '
    'để tiếp tục đà tăng Hiển Thị (+6,7%) và Sự kiện QT (+80%).', label_color='375623', good=True)

r += 1
write_bullet(r, '  [Ưu tiên 4]',
    'Tối ưu Landing Page: nhân rộng mô hình đang hiệu quả (SQ +75%) sang các landing page '
    'khác, A/B test tiêu đề và form đăng ký.', label_color='375623', good=True)

ws.row_dimensions[r].height = 22

wb.save(str(src))
print('Done! Đã thêm phần nhận xét vào sheet "Báo Cáo T4 vs T5".')
